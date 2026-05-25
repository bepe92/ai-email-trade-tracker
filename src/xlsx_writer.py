"""Write parsed deals to a styled .xlsx workbook with two sheets: Deals / To Review."""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEALS_HEADERS = [
    ("File", "file", 28),
    ("Broker", "broker", 26),
    ("Product", "product", 32),
    ("Volume (MT)", "volume_mt", 14),
    ("Price (USD)", "price_usd", 14),
    ("Price Unit", "price_unit", 18),
    ("Trade Date", "trade_date", 14),
    ("Reference", "reference", 18),
]

REVIEW_HEADERS = [
    ("File", "file", 32),
    ("Errors", "errors", 50),
    ("Broker", "broker", 26),
    ("Product", "product", 28),
    ("Volume (MT)", "volume_mt", 14),
    ("Price (USD)", "price_usd", 14),
    ("Trade Date", "trade_date", 14),
    ("Reference", "reference", 18),
]

HEADER_FILL = PatternFill("solid", fgColor="003366")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
REVIEW_FILL = PatternFill("solid", fgColor="B30000")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALT_FILL = PatternFill("solid", fgColor="F4F7FB")


def write_workbook(out_path: Path, successes: list[dict], reviews: list[dict]):
    wb = Workbook()
    _write_deals_sheet(wb.active, successes)
    _write_review_sheet(wb.create_sheet("To Review"), reviews)
    wb.save(out_path)


def _write_deals_sheet(ws, rows: list[dict]):
    ws.title = "Deals"
    _write_headers(ws, DEALS_HEADERS, HEADER_FILL)

    for i, row in enumerate(rows, start=2):
        fill = ALT_FILL if i % 2 == 0 else None
        for col_idx, (_, key, _) in enumerate(DEALS_HEADERS, start=1):
            value = row.get(key)
            if key == "trade_date" and value:
                try:
                    value = datetime.strptime(value, "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if key in ("volume_mt",):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif key in ("price_usd",):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif key == "trade_date":
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_review_sheet(ws, rows: list[dict]):
    _write_headers(ws, REVIEW_HEADERS, REVIEW_FILL)

    for i, row in enumerate(rows, start=2):
        parsed = row.get("parsed") or {}
        errors_text = "\n".join(row.get("errors", []))
        values = {
            "file": row.get("file"),
            "errors": errors_text,
            "broker": parsed.get("broker"),
            "product": parsed.get("product"),
            "volume_mt": parsed.get("volume_mt"),
            "price_usd": parsed.get("price_usd"),
            "trade_date": parsed.get("trade_date"),
            "reference": parsed.get("reference"),
        }
        for col_idx, (_, key, _) in enumerate(REVIEW_HEADERS, start=1):
            cell = ws.cell(row=i, column=col_idx, value=values.get(key))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(key == "errors"))

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def _write_headers(ws, headers, fill):
    for col_idx, (label, _, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
